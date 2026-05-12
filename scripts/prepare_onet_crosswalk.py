"""Maintainer script: ESCO <-> O*NET crosswalk xlsx -> bundled crosswalk.csv.

Downloads the ESCO/O*NET crosswalk from the O*NET Resource Center as an
xlsx, filters it to the curated ESCO occupation subset produced by
``prepare_esco.py``, joins each row to its ESCO occupation URI via the ESCO
``code`` column, and writes ``src/career_planner/data/crosswalk.csv``.

The output CSV starts with attribution and provenance comment lines, then a
header row, then the data rows. The bundled loader should skip lines that
start with ``#`` before parsing the header row.

If the upstream URL has moved (O*NET reorganizes downloads periodically),
drop the file at ``scripts/raw/onet/ESCO_to_ONET-SOC.xlsx`` and re-run.

Run with: ``python scripts/prepare_onet_crosswalk.py``
"""

from __future__ import annotations

import csv
import io
import sys
from datetime import date
from pathlib import Path

import httpx
import openpyxl
import yaml
from rich.console import Console

REPO_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = REPO_ROOT / "scripts" / "raw" / "onet"
OUT_DIR = REPO_ROOT / "src" / "career_planner" / "data"
OUT_FILE = OUT_DIR / "crosswalk.csv"
LOCAL_FALLBACK = RAW_DIR / "ESCO_to_ONET-SOC.xlsx"

# Upstream download. Update if O*NET reorganizes its file layout.
DEFAULT_URL = "https://www.onetcenter.org/crosswalks/esco/ESCO_to_ONET-SOC.xlsx"

HEADER_LABEL = "ESCO/ISCO Code"
OUTPUT_COLUMNS = (
    "esco_uri",
    "esco_code",
    "esco_title",
    "onet_soc_code",
    "onet_soc_title",
)

console = Console()


def _curated_code_to_uri() -> dict[str, str]:
    path = OUT_DIR / "esco-occupations.yml"
    if not path.is_file():
        console.print(
            f"[red]{path.relative_to(REPO_ROOT)} not found.[/red] "
            "Run scripts/prepare_esco.py first."
        )
        sys.exit(1)
    with path.open("r", encoding="utf-8") as fh:
        data = yaml.safe_load(fh) or {}
    mapping: dict[str, str] = {}
    for row in data.get("occupations", []):
        code = (row.get("code") or "").strip()
        uri = (row.get("uri") or "").strip()
        if code and uri:
            mapping[code] = uri
    return mapping


def _download(url: str) -> bytes:
    console.print(f"Downloading [cyan]{url}[/cyan]")
    with httpx.Client(timeout=120.0, follow_redirects=True) as client:
        with client.stream("GET", url) as resp:
            resp.raise_for_status()
            total = int(resp.headers.get("content-length") or 0)
            chunks: list[bytes] = []
            with console.status("  fetching...") as status:
                received = 0
                for chunk in resp.iter_bytes():
                    chunks.append(chunk)
                    received += len(chunk)
                    if total:
                        pct = 100 * received / total
                        status.update(
                            f"  fetched {received:,}/{total:,} bytes ({pct:.1f}%)"
                        )
                    else:
                        status.update(f"  fetched {received:,} bytes")
            return b"".join(chunks)


def _resolve_xlsx_bytes() -> bytes:
    if LOCAL_FALLBACK.is_file():
        console.print(f"Using local file [cyan]{LOCAL_FALLBACK}[/cyan]")
        return LOCAL_FALLBACK.read_bytes()
    try:
        return _download(DEFAULT_URL)
    except httpx.HTTPError as exc:
        console.print(f"[red]Download failed:[/red] {exc}")
        console.print(
            "Manually download the ESCO/O*NET crosswalk from\n"
            "  https://www.onetcenter.org/crosswalks.html\n"
            f"and place the xlsx at {LOCAL_FALLBACK}."
        )
        sys.exit(1)


def _iter_xlsx_rows(payload: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_seen = False
    for row in ws.iter_rows(values_only=True):
        if not header_seen:
            if row and row[0] == HEADER_LABEL:
                header_seen = True
            continue
        if row is None or all(cell is None for cell in row):
            continue
        yield row


def _filter_rows(
    payload: bytes, code_to_uri: dict[str, str]
) -> list[tuple[str, str, str, str, str]]:
    out: list[tuple[str, str, str, str, str]] = []
    seen_codes: set[str] = set()
    for row in _iter_xlsx_rows(payload):
        esco_code = (row[0] or "").strip() if len(row) > 0 else ""
        esco_title = (row[1] or "").strip() if len(row) > 1 else ""
        onet_code = (row[2] or "").strip() if len(row) > 2 else ""
        onet_title = (row[3] or "").strip() if len(row) > 3 else ""
        uri = code_to_uri.get(esco_code)
        if not uri or not onet_code:
            continue
        seen_codes.add(esco_code)
        out.append((uri, esco_code, esco_title, onet_code, onet_title))
    out.sort(key=lambda r: (r[1], r[3]))
    return out


def main() -> None:
    today = date.today().isoformat()
    console.rule("[bold]Preparing ESCO <-> O*NET crosswalk")

    code_to_uri = _curated_code_to_uri()
    console.print(f"Curated ESCO occupations: {len(code_to_uri):,}")

    payload = _resolve_xlsx_bytes()
    rows = _filter_rows(payload, code_to_uri)
    matched_codes = {row[1] for row in rows}
    console.print(f"Crosswalk rows kept:       {len(rows):,}")
    console.print(
        f"ESCO occupations covered:  {len(matched_codes):,} / "
        f"{len(code_to_uri):,}"
    )

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with OUT_FILE.open("w", encoding="utf-8", newline="") as fh:
        fh.write(
            "# Sources: ESCO v1.2.1 (European Commission) + "
            "O*NET 29.0 (USDOL/ETA)\n"
        )
        fh.write("# See THIRD_PARTY_NOTICES.md.\n")
        fh.write(f"# version_date: {today}\n")
        fh.write(
            "# source: O*NET Resource Center ESCO crosswalk "
            "(https://www.onetcenter.org/crosswalks.html)\n"
        )
        writer = csv.writer(fh)
        writer.writerow(OUTPUT_COLUMNS)
        writer.writerows(rows)

    size_kb = OUT_FILE.stat().st_size / 1024
    console.print(f"  wrote {OUT_FILE.relative_to(REPO_ROOT)}  ({size_kb:,.1f} KB)")


if __name__ == "__main__":
    main()
